import cv2
import os
import shutil
import pytesseract
from matplotlib import pyplot as pt
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.borders import Border, Side
from drawingNum import GetString
from table_parser import TableParser, TableGrid

# --- Tesseract path: auto-detect or set via environment variable --- #
_tess_env = os.environ.get("TESSERACT_CMD")
if _tess_env:
    pytesseract.pytesseract.tesseract_cmd = _tess_env
elif shutil.which("tesseract"):
    pytesseract.pytesseract.tesseract_cmd = shutil.which("tesseract")
else:
    # Fallback for Windows default install location
    pytesseract.pytesseract.tesseract_cmd = (
        r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    )

KEYWORDS = [
    "DRAWING NUMBER", "DRAWING NO",
    "DRAWN BY", "DRAWN",
    "CHECKED BY", "CHECKED",
    "TITLE", "DRAWING TITLE",
    "APPPROVED BY", "APPROVED",
    "CONTRACTOR", "COMPANY",
    "UNIT", "STATUS", "PAGE",
    "PROJECT NO", "PROJECT NUM",
    "LANG", "CAD NO",
    "FONT", "FONT STYLE",
    "AMENDMENTS",
]

# Pairs where the shorter variant is redundant when the longer one is present
REDUNDANT_PAIRS = [
    ("CHECKED", "CHECKED BY"),
    ("DRAWN", "DRAWN BY"),
    ("TITLE", "DRAWING TITLE"),
    ("APPROVED", "APPROVED BY"),
]

parser = TableParser()
wb = Workbook()

for image in range(1, 21):
    img_path = os.path.join("images", f'{image:02}.png')

    if not os.path.exists(img_path):
        print("Image ", f'{image:02}.png', " not found.")
        continue

    init_img = cv2.imread(img_path, 0)
    [init_row, init_col] = init_img.shape


    # --- Cropping border --- #
    img = init_img[12:init_row - 15, 12:init_col - 12]
    [nrow, ncol] = img.shape


    # --- Isolating vertical & horizontal lines --- #
    _, bin_img = cv2.threshold(img, 127, 255, cv2.THRESH_BINARY_INV)

    horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, ncol // 150))
    eroded_verti = cv2.erode(bin_img, horiz_kernel, iterations=5)
    vertical_lines = cv2.dilate(eroded_verti, horiz_kernel, iterations=5)

    verti_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (nrow // 150, 1))
    eroded_hori = cv2.erode(bin_img, verti_kernel, iterations=5)
    horizontal_lines = cv2.dilate(eroded_hori, verti_kernel, iterations=5)

    combined_lines = cv2.bitwise_or(vertical_lines, horizontal_lines)


    # --- Drawing removal mask --- #
    rect_kernel3 = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    drawingMask = cv2.erode(combined_lines, rect_kernel3, iterations=2)
    drawingMask = cv2.dilate(drawingMask, rect_kernel3, iterations=50)
    table_lines = drawingMask + np.bitwise_not(combined_lines)


    # --- Removing small noise / arrow contours --- #
    table_lines_dil = cv2.dilate(np.bitwise_not(table_lines), rect_kernel3,
                                 iterations=5)

    contours, hierarchy = cv2.findContours(table_lines_dil, cv2.RETR_TREE,
                                           cv2.CHAIN_APPROX_SIMPLE)
    sorted_contours = sorted(contours, key=cv2.contourArea, reverse=False)

    table_bgr = cv2.cvtColor(table_lines, cv2.COLOR_GRAY2BGR)
    for i, cntr in enumerate(sorted_contours):
        x, y, w, h = cv2.boundingRect(cntr)
        if w < 30 or h < 30:
            cv2.drawContours(table_bgr, sorted_contours, i, (255, 255, 255),
                             thickness=-1)

    table_only = cv2.cvtColor(table_bgr, cv2.COLOR_BGR2GRAY)
    _, table_only = cv2.threshold(table_only, 150, 255, cv2.THRESH_BINARY)


    # --- Isolating table cells --- #
    table_only_copy = cv2.copyMakeBorder(table_only, 5, 5, 5, 5,
                                         cv2.BORDER_CONSTANT, 0)
    table_lines_dil2 = cv2.dilate(np.bitwise_not(table_only_copy), rect_kernel3,
                                  iterations=1)
    cell_cntr, hierarchy = cv2.findContours(table_lines_dil2, cv2.RETR_TREE,
                                            cv2.CHAIN_APPROX_SIMPLE)


    # ------------------------------------------------------------------
    # NEW: structured grid-based table parsing via TableParser
    # ------------------------------------------------------------------
    # Build a binary image of just the table lines (white lines on black bg)
    table_bin = np.bitwise_not(table_only)  # lines are white
    grid = parser.parse_table_region(img, binary_lines=table_bin)

    # Use grid-based extraction when the parser found a valid structure
    grid_results = []
    if grid.num_rows >= 1 and grid.num_cols >= 1:
        grid_results = parser.extract_labeled_values(grid, KEYWORDS)

    # ------------------------------------------------------------------
    # LEGACY: contour-based keyword matching (fallback / merge)
    # ------------------------------------------------------------------
    table_bgr2 = cv2.cvtColor(table_only, cv2.COLOR_GRAY2BGR)
    useful_cells = []

    for c in cell_cntr:
        coordinates = cv2.boundingRect(c)
        x, y, w, h = coordinates
        rect_area = w * h
        if rect_area < ((nrow // 4) * (ncol // 4)) and h < 400:
            cell = img[y:y + h, x:x + w]
            string = pytesseract.image_to_string(cell,
                                                 config='--psm 6').strip()
            string_list = string.splitlines()
            for k in KEYWORDS:
                if k in string:
                    useful_cells.append([k, coordinates, string_list])

            cv2.rectangle(table_bgr2, (x, y), (x + w, y + h), (0, 0, 0), -1)

    table_mask = cv2.cvtColor(table_bgr2, cv2.COLOR_BGR2GRAY)
    table_mask = cv2.dilate(np.bitwise_not(table_mask), rect_kernel3,
                            iterations=5)

    drawing = np.bitwise_not(bin_img) + table_mask
    drawing[drawing >= 5] = 255
    drawing[drawing < 5] = 0

    tables = np.bitwise_not(bin_img) + np.bitwise_not(table_mask)
    tables[tables >= 5] = 255
    tables[tables < 5] = 0


    # --- Fallback: full-vertical table detection --- #
    _, bin_drawing = cv2.threshold(drawing, 150, 255, cv2.THRESH_BINARY_INV)
    bin_drawing = cv2.erode(bin_drawing, horiz_kernel, iterations=5)
    bin_drawing = cv2.dilate(bin_drawing, horiz_kernel, iterations=5)

    vertical_lines_dil = cv2.dilate(bin_drawing, rect_kernel3, iterations=2)
    vert_contours, _ = cv2.findContours(vertical_lines_dil, cv2.RETR_TREE,
                                        cv2.CHAIN_APPROX_SIMPLE)
    vert_tf = any(
        cv2.boundingRect(c)[3] >= nrow - 50 for c in vert_contours
    )

    if vert_tf and len(useful_cells) == 0 and not grid_results:
        drawing_mask2 = np.zeros((nrow, ncol), dtype=np.uint8)
        contours2, _ = cv2.findContours(np.bitwise_not(bin_img), cv2.RETR_TREE,
                                        cv2.CHAIN_APPROX_NONE)
        sorted_contours2 = sorted(contours2, key=cv2.contourArea, reverse=True)
        x, y, w, h = cv2.boundingRect(sorted_contours2[0])
        drawing_mask2[y:y + h, x:x + w] = 255

        tables = np.bitwise_not(bin_img) + drawing_mask2
        tables[tables >= 5] = 255
        tables[tables < 5] = 0

        drawing = np.bitwise_not(bin_img) + np.bitwise_not(drawing_mask2)
        drawing[drawing >= 5] = 255
        drawing[drawing < 5] = 0

        drawingNum = GetString(init_img, "DRAWING NUMBER", "DRAWING NO")
        drawnBy = GetString(init_img, "DRAWN BY", "DRAWN")
        if drawingNum and len(drawingNum) > 0:
            useful_cells.append(["DRAWING NUMBER", None, ["", drawingNum]])


    # --- Re-analyze cells that contain only the label (no value) --- #
    for index, info in enumerate(useful_cells):
        if len("".join(info[2])) < len(info[0]) + 3:
            x, y, w, h = info[1]
            y_range = y + h + (300 if info[0] == "AMENDMENTS" else 80)
            y_range = min(y_range, nrow)
            cell = img[y:y_range, x:x + w]

            _, cell_thresh = cv2.threshold(cell, 150, 255,
                                           cv2.THRESH_BINARY_INV)
            cell_clean = parser.clean_cell_for_ocr(
                img[y:y_range, x:x + w]
            )
            string = pytesseract.image_to_string(cell_clean,
                                                 config='--psm 6').strip()
            info[2] = string.splitlines()
            useful_cells[index] = info


    # ------------------------------------------------------------------
    # Merge grid-based results with legacy results
    # ------------------------------------------------------------------
    # Prefer grid-based results; fill gaps from legacy
    seen_keywords = set()
    merged_data = []

    for gr in grid_results:
        kw = gr["keyword"]
        seen_keywords.add(kw)
        # Build the same [keyword, coords, [label, value]] structure
        cell = gr["cell"]
        merged_data.append([kw, (cell.x, cell.y, cell.w, cell.h),
                            [kw, gr["value"]]])

    for legacy in useful_cells:
        if legacy[0] not in seen_keywords:
            seen_keywords.add(legacy[0])
            merged_data.append(legacy)


    # ------------------------------------------------------------------
    # Deduplication & cleanup
    # ------------------------------------------------------------------
    table_data = []
    for c in merged_data:
        if c not in table_data:
            table_data.append(c)

    table_data.sort(key=lambda lst: lst[0])

    # Pop amendments table data
    def _indices_of(word, data):
        return [i for i, e in enumerate(data) if e[0] == word]

    amend_index = _indices_of("AMENDMENTS", table_data)
    amendments = None

    if len(amend_index) == 1:
        amendments = table_data.pop(amend_index[0])
    elif len(amend_index) > 1:
        removed = 0
        for i in amend_index:
            idx = i - removed
            if len(table_data[idx][2]) > 1 and len(table_data[idx][2][1]) > 3:
                amendments = table_data.pop(idx)
            else:
                del table_data[idx]
            removed += 1

    # Remove shorter keyword when a longer variant is present
    for redundant, keep in REDUNDANT_PAIRS:
        kws = [c[0] for c in table_data]
        if keep in kws and redundant in kws:
            del table_data[kws.index(redundant)]


    # ------------------------------------------------------------------
    # NEW: Parse amendments as a structured sub-table when possible
    # ------------------------------------------------------------------
    amendments_rows = []
    amendments_headers = []

    # Try grid-based sub-table extraction first
    if grid.num_rows >= 1 and grid.num_cols >= 1:
        sub = parser.parse_sub_table(grid, "AMENDMENTS")
        if sub and sub.num_rows >= 1:
            nested = sub.to_nested_dict()
            if nested:
                amendments_headers = list(nested[0].keys())
                amendments_rows = [
                    [row.get(h, "") for h in amendments_headers]
                    for row in nested
                ]

    # Fall back to legacy string-split approach
    if not amendments_rows and amendments is not None:
        if len(amendments[2]) >= 2:
            title_parts = amendments[2][1].split()
            if title_parts:
                amendments_headers = title_parts
                for line_idx in range(2, len(amendments[2])):
                    row_parts = amendments[2][line_idx].split()
                    if row_parts:
                        # Pad or truncate to match header count
                        while len(row_parts) < len(amendments_headers):
                            row_parts.append("")
                        amendments_rows.append(
                            row_parts[:len(amendments_headers)]
                        )


    # --- Writing table data into .xlsx file --- #
    ws = wb.create_sheet(f'{image:02}.png', image)
    ws.append(["Field Title", "Content"])

    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick'),
    )
    ws.cell(row=1, column=1).border = thick_border
    ws.cell(row=1, column=2).border = thick_border

    for info in table_data:
        value = info[2][1] if len(info[2]) > 1 else ""
        ws.append([info[0], value])

    # Write amendments sub-table
    if amendments_headers:
        letters = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        n_cols = min(len(amendments_headers), len(letters))
        ws['E1'] = "Amendments"

        # Header row
        for i in range(n_cols):
            ws[f'{letters[i]}2'] = amendments_headers[i]

        # Data rows
        for r_idx, row in enumerate(amendments_rows):
            for c_idx in range(n_cols):
                ws[f'{letters[c_idx]}{3 + r_idx}'] = (
                    row[c_idx] if c_idx < len(row) else ""
                )

        end_row = 2 + len(amendments_rows)
        end_col_letter = letters[n_cols - 1]
        tab = Table(
            displayName=f"Amendments{image}",
            ref=f"E2:{end_col_letter}{end_row}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    wb.save(filename='drawingInfo.xlsx')

    # Save extracted drawing image
    writeFolder = "extracted"
    if not os.path.exists(writeFolder):
        os.makedirs(writeFolder)

    output_image_path = os.path.join(writeFolder, f'drawing{image:02}.png')
    cv2.imwrite(output_image_path, drawing)

    print("Image ", f'{image:02}.png', " processed.")
